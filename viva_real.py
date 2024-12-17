# Pesquisa no site da viva real. Vai ter uma planilha de input, com o nome do bairro e da cidade, com o tipo de contrato(aluguel ou compra) 
# e você tem que capturar: Valor, Area, Cresci, Imobiliaria e telefone.
# Depois de pesquisar, tem que ir anuncio em anuncio e pegando as fotos e criar um PDF com as Fotos e tirar um print da pagina do anuncio. 
# Precisa pegar a data e a hora que foi feita a pesquisa e o link de cada resultado que retornou


from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from time import sleep
from openpyxl import load_workbook
import undetected_chromedriver as uc
import requests
import os
from fpdf import FPDF
from PIL import Image

# Abrindo o navegador
navegador = uc.Chrome()
navegador.maximize_window()
sleep(2)

# Carregando a planilha
planilha = 'Viva_Real.xlsx'
abrir_planilha = load_workbook(planilha)
sheet = abrir_planilha['Planilha1']

# Percorrendo as linhas
for linha in sheet.iter_rows(2, sheet.max_row, values_only=True):
    nome_bairro, nome_cidade, tipo_contrato = linha
           
    # Criando as variáveis para comparar com a planilha e acessar o site baseado no resultado da condição
    contrato_alugar = "Alugar"
    contrato_comprar = "Comprar"

    # Comparando contrato da planilha com o do site
    if tipo_contrato == contrato_alugar:
        navegador.get("https://www.vivareal.com.br/aluguel/")
        sleep(3)
    else:
        navegador.get("https://www.vivareal.com.br/venda/")
        sleep(3)

    #Localizando campo de pesquisa e preenchendo.
    preencher = navegador.find_element(By.XPATH, './/input[@id="filter-location-search-input"]')
    preencher.clear()
    preencher.send_keys(f"{nome_bairro} {nome_cidade}")
    sleep(5)
    preencher.send_keys(Keys.ENTER)
    sleep(3)
    
    # Criando variável que vai receber lista de imoveis
    lista_imoveis = []
    
    # Localizando elemento pai
    imoveis = navegador.find_elements(By.XPATH, './/div[@data-type="property"]')
    
    # Criando a classe PDF que vai colocar as imagens dentro dele
    class PDF(FPDF):
            def header(self):
                self.set_font('Arial', 'B', 12)
                self.cell(0,10, 'Imagens dos Anúncios', 0, 1, 'C')
                self.ln(10)
    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Contador para teste
    contador = 0
    # Criando o loop que vai utilizar o elemento pai de base, onde vai passar anuncio por anuncio    
    for imovel in imoveis:
        if contador>=2:
            break
        
        # Setando a aba da pesquisa como principal
        aba_principal = navegador.current_window_handle
        
        #Clicando no anuncio que abrirá em outra aba
        imovel.find_element(By.XPATH, './/div[@class="property-card__content"]').click()
        sleep(4)
        
        # Setando a lista de abas do navegador para a variável abas
        abas = navegador.window_handles
        
        # Mudando o foco para a última aba aberta. (colocando -1 sempre vai buscar a ultima aba aberta)
        navegador.switch_to.window(abas[-1])
        sleep(1)
        
        try:
            #Capturando o valor do imóvel
            valor_imovel = WebDriverWait(navegador,5).until(EC.visibility_of_element_located(((By.XPATH, './/p[@data-testid="price-info-value"]')))).text
            sleep(1)
        except:
            print("Elemento não localizado")
            break
        
        # Capturando a área do imóvel
        area_imovel = navegador.find_element(By.XPATH, './/p[@itemprop="floorSize"]/span[@data-cy="ldp-propertyFeatures-txt"]').text
        sleep(1)
        
        # Capturando a imobiliária
        imobiliaria = navegador.find_element(By.XPATH, '(//div/a[@title="Loja Oficial do Anunciante"])[2]').text
        sleep(1)
        
        # Capturando a URL da página atual
        url = navegador.current_url  
        sleep(1)
        
        try:
            cresci = WebDriverWait(navegador, 5).until(EC.visibility_of_element_located((
                By.XPATH, '(.//div[@class="advertiser-header__infos-wrapper"]/p/text()[2])[2]'))).text
        
        except :
            cresci = "sem creci"
        sleep(1)
        
        
        # Localizando e clicando o elemento do botão que expõe o(s) telefone(s)
        botao_ver_telefone = navegador.find_element(By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[2]/section/div[3]/button').click()
        sleep(2)

        # Ao clicar pra visualizar o telefone a primeira vez, é solicitado algumas infos, então coloquei pra checar se o elemento
        # de cadastro de nome está visível. Se estiver, vai seguir o fluxo do cadastro, se não vai continuar o fluxo normal
        try:
            cadastro_nome = WebDriverWait(navegador, 3).until(EC.visibility_of_element_located((
                By.XPATH, '(.//input[@placeholder="Nome"])[3]')))
            cadastro_nome.send_keys("Teste")
            
            cadastro_telefone = navegador.find_element(By.XPATH, '(.//input[@placeholder="Telefone"])[3]')
            cadastro_telefone.send_keys("3525817339")
            
            cadastro_email = navegador.find_element(By.XPATH, '(.//input[@placeholder="E-mail"])[3]')
            cadastro_email.send_keys("teste@teste.com")
            
            botao_submit = navegador.find_element(By.XPATH, '(.//button[@type="submit"])[3]').click()
            sleep(2)
        except:
            print("Elemento não encontrado, seguindo o fluxo...")      
               
        # Captura o primeiro telefone
        telefone1 = WebDriverWait(navegador, 5).until(EC.visibility_of_element_located((By.XPATH, './/a[@data-cy="lead-modalPhone-phonesList-txt"][1]'))).text
        sleep(1)
        
        # Alguns anúncios tem dois telefones, então vai aguardar 5 segundos e verificar se o elemento do segundo telefone está na tela,
        # se estiver ele captura, se não estiver a variável receberá "Sem Telefone"
        try:
            telefone2 = WebDriverWait(navegador, 5).until(EC.visibility_of_element_located((By.XPATH, '(.//a[@data-cy="lead-modalPhone-phonesList-txt"])[2]'))).text
        except:
            telefone2 = "Sem telefone"
        sleep(1)
        
        # Fechando a tela de telefones
        fechar_tela_telefone = navegador.find_element(By.XPATH, './/span[@data-testid="on-close"]')
        fechar_tela_telefone.click()
        sleep(1)    
           
        # Clicando no botão das fotos
        botao_fotos = navegador.find_element(By.XPATH, '(.//button[@type="button"])[1]')
        botao_fotos.click()
        
        
        # Capturando as imagens
        imagens_elementos = WebDriverWait(navegador, 5).until(EC.presence_of_all_elements_located((
            By.XPATH, './/div[@class="image-container__item"]/picture/source[@type="image/webp"]')))
        imagens_urls = [img.get_attribute("srcset") for img in imagens_elementos]
        
        def is_valid_image(img_path):
            try:
                with Image.open(img_path) as img:
                    img.verify()  # Valida a integridade da imagem
                return True
            except Exception as e:
                print(f"Imagem inválida: {img_path}, erro: {e}")
                return False
    
        # Criar diretório pra salvar as imagens
        os.makedirs("imagens", exist_ok=True)
        
        # Listando e baixando as imagens
        imagens_salvas = []
        for i, url in enumerate(imagens_urls):
            if not url:  # Ignorar URLs vazias
                continue
            try:
                response = requests.get(url, timeout=10)
                if response.status_code == 200:
                    img_path = f"imagens/imagem_{i}.jpg"
                    with open(img_path, "wb") as f:
                        f.write(response.content)
                    
                    # Validar se o arquivo é uma imagem válida
                    if is_valid_image(img_path):
                        imagens_salvas.append(img_path)
                    else:
                        os.remove(img_path)  # Remover arquivos inválidos
                else:
                    print(f"Falha ao baixar a imagem {url}: Status {response.status_code}")
            except requests.exceptions.RequestException as e:
                print(f"Erro ao baixar a imagem {url}: {e}")
        


        # Adicionar as imagens no PDF
        for img_path in imagens_salvas:
            try:
                pdf.set_font("Arial", size=12)
                pdf.cell(0, 10, f"Imagem: {os.path.basename(img_path)}", ln=True)
                pdf.image(img_path, x=10, y=pdf.get_y(), w=100)
                pdf.ln(60)
            except RuntimeError as e:
                print(f"Erro ao adicionar a imagem {img_path} ao PDF: {e}")

        # Salvar o PDF
        pdf.output("imagens_anuncios.pdf")
        print(f"PDF gerado com sucesso: imagens_anuncios.pdf")


        print(f"Tipo de contrato: {tipo_contrato}")
        print(f"Valor do imóvel: {valor_imovel}")
        print(f"Área do imóvel: {area_imovel}")
        print(f"Creci: {cresci}")
        print(f"Imobiliária: {imobiliaria}")
        print(f"Telefone 1: {telefone1}")
        print(f"Telefone 2: {telefone2}")
        print(f"URL da página: {url}")
        print("-" * 120)
        
        navegador.close()
        sleep(2)
        navegador.switch_to.window(aba_principal)
        sleep(2)

        contador += 1

   # for imovel in lista_imoveis:
   #     print(f"{imovel['valor_imovel']:<10}{imovel['area_imovel']:<5}{imovel['cresci']:<10}{imovel['imobiliaria']:<20}{imovel['telefone1']:<10}{imovel['telefone2']:<10}{imovel['url']:<50}")
   #     print("-" * 70)
 
   
input("Enter para encerrar...")