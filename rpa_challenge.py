from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from time import sleep

# atribuindo planilha a variável (se estiver em outra pasta deve-se colocar o caminho completo da pasta)
planilha = 'challenge.xlsx' 

# atribuindo o carregamento da planilha a uma variável
abrir_planilha = load_workbook(planilha) 

# selecionando a sheet a partir da planilha carregada no load_workbook
sheet_selecionada = abrir_planilha['Sheet1'] 

# atribuindo o controle do navegador a variável e abrindo o mesmo, toda ação realizada no navegador passará por esta variável
abrir_navegador = webdriver.Chrome() 
sleep(1)

# acessando a URL
abrir_navegador.get("https://rpachallenge.com/") 
sleep(1)

# maximizando a janela
abrir_navegador.maximize_window() 
sleep(1)

# clicando no botão de start a partir do XPATH
start = abrir_navegador.find_element(By.XPATH, "/html/body/app-root/div[2]/app-rpa1/div/div[1]/div[6]/button") 
start.click()

# Percorrendo a planilha através do for. (como a planilha tem cabeçalho, começa-se a percorrer a planilha pela linha 2, sendo necessário colocar um +1 no final para compensar)
for linha in range(2, len(sheet_selecionada['A']) + 1):
    first_name = sheet_selecionada[f'A{linha}'].value
    last_name = sheet_selecionada[f'B{linha}'].value
    company_name = sheet_selecionada[f'C{linha}'].value
    role_company = sheet_selecionada[f'D{linha}'].value
    adress = sheet_selecionada[f'E{linha}'].value
    email = sheet_selecionada[f'F{linha}'].value
    phone_number = sheet_selecionada[f'G{linha}'].value

    # Preenchimento dos campos da página.
    # A planilha tem campos que estão vazios no final, então só serão preenchidos campos onde a célula está preenchida
    if first_name != None: 
        # Seleciona o campo de input onde será preenchido, procurando no HTML um elemento <input> que tenha o atributo ng-reflect-name='labelFirstName'
        preencher = abrir_navegador.find_element(By.XPATH, "//input[@ng-reflect-name='labelFirstName']") 

        # Limpa o campo caso esteja com algum caracter
        preencher.clear() 

        # O método 'send keys' realiza a simulação de digitação de texto, nesse caso o valor que foi pego no 'first_name' é "digitado" no campo
        preencher.send_keys(first_name) 
        

        preencher = abrir_navegador.find_element(By.XPATH, "//input[@ng-reflect-name='labelLastName']")
        preencher.clear()
        preencher.send_keys(last_name)
        

        preencher = abrir_navegador.find_element(By.XPATH, "//input[@ng-reflect-name='labelCompanyName']")
        preencher.clear()
        preencher.send_keys(company_name)
        

        preencher = abrir_navegador.find_element(By.XPATH, "//input[@ng-reflect-name='labelRole']")
        preencher.clear()
        preencher.send_keys(role_company)
        

        preencher = abrir_navegador.find_element(By.XPATH, "//input[@ng-reflect-name='labelAddress']")
        preencher.clear()
        preencher.send_keys(adress)
        

        preencher = abrir_navegador.find_element(By.XPATH, "//input[@ng-reflect-name='labelEmail']")
        preencher.clear()
        preencher.send_keys(email)
        

        preencher = abrir_navegador.find_element(By.XPATH, "//input[@ng-reflect-name='labelPhone']")
        preencher.clear()
        preencher.send_keys(phone_number)
        
        # clica no botão 'Submit' a cada rodada do preenchimento dos campos
        abrir_navegador.find_element(By.XPATH, "//input[@value='Submit']").click() 
        



input('Enter')