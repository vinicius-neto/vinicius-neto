

# Esta é uma versão mais dinâmica da primeira, que acessava diretamente célula por célula.
# A ideia desta versão é utilizar mais as funções do openpyxl invés de acessar diretamente as células

import undetected_chromedriver as uc
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from time import sleep
from openpyxl import load_workbook

#Carregando planilha
caminho_planilha = "challenge.xlsx"
carregar_planilha = load_workbook(caminho_planilha)
sheet = carregar_planilha['Sheet1']

# Abrindo o navegador
navegador = uc.Chrome()
navegador.get("https://rpachallenge.com/")
navegador.maximize_window()
sleep(2)

#Iterando sobre as linhas
for linha in sheet.iter_rows(2, sheet.max_row, values_only=True):
    first_name, last_name, company_name, role_company, adress, email, phone_number = linha
    
    try:
        #Startando o processo
        botao_start = WebDriverWait(navegador, 5).until(ec.visibility_of_element_located((By.XPATH, './/button[@_ngcontent-c1]')))
        botao_start.click()
    except:
        print("Botão de start não encontrado")

    #Preenchendo os campos
    preencher = navegador.find_element(By.XPATH, './/input[@ng-reflect-name="labelFirstName"]')
    preencher.send_keys(first_name)
    

    preencher = navegador.find_element(By.XPATH, './/input[@ng-reflect-name="labelLastName"]')
    preencher.send_keys(last_name)
    

    preencher = navegador.find_element(By.XPATH, './/input[@ng-reflect-name="labelCompanyName"]')
    preencher.send_keys(company_name)
    

    preencher = navegador.find_element(By.XPATH, './/input[@ng-reflect-name="labelRole"]')
    preencher.send_keys(role_company)
    

    preencher = navegador.find_element(By.XPATH, './/input[@ng-reflect-name="labelAddress"]')
    preencher.send_keys(adress)
    

    preencher = navegador.find_element(By.XPATH, './/input[@ng-reflect-name="labelEmail"]')
    preencher.send_keys(email)
    

    preencher = navegador.find_element(By.XPATH, './/input[@ng-reflect-name="labelPhone"]')
    preencher.send_keys(phone_number)
    
    #Clicando em submit
    submit = navegador.find_element(By.XPATH, './/input[@type="submit"]').click()
    sleep(0.2)

input("Enter para encerrar...")
